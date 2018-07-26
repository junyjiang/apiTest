# -*- coding: utf-8 -*-
import openpyxl
from .Log import LOG
#@logger('解析测试用例文件')
def datacel(casename):
    try:
         filepath='../test_Data/testcase.xlsx'
         file = openpyxl.load_workbook(filepath)
         #openpyxl更新到2.5后get_sheet_names@deprecated("Use wb.sheetnames")
         sheets_names = file.sheetnames
         listid=[]
         listparem=[]
         listurl=[]
         listfangshi=[]
         listqiwang=[]
         listjsonkey=[]
         listname=[]
         listSql=[]
         for i in range(len(sheets_names)):
             # openpyxl更新到2.5后get_sheet_by_name---- @deprecated("Use wb[sheetname]")
             sheet = file[sheets_names[i]]
             rows = sheet.max_row + 1
             if sheets_names[i] == casename:
                 for j in range(2, rows):
                     listid.append(sheet["A" + str(j)].value)
                     listname.append(sheet["B" + str(j)].value)
                     listparem.append(sheet["C" + str(j)].value)
                     listurl.append(sheet["D" + str(j)].value)
                     listfangshi.append(sheet["E" + str(j)].value)
                     listqiwang.append(sheet["F" + str(j)].value)
                     listjsonkey.append(sheet["G" + str(j)].value)
                     listSql.append(sheet["H" + str(j)].value)
                 return listid, listparem, listurl, listfangshi, listqiwang, listname, listjsonkey, listSql
    except:LOG.info('打开测试用例失败，原因是:%s'%Exception)
#@logger('生成数据驱动所用数据')
def makedata(casename):
    listid, listparem, listurl, listfangshi, listqiwang, listname,listjsonkey,listSql = datacel(casename)
    make_data = []
    for i in range(len(listid)):
        make_data.append({'name':listname[i],'url': listurl[i], 'parem': listparem[i], 'fangshi': listfangshi[i],
                          'qiwang': listqiwang[i],'jsonkey':listjsonkey[i],'listSql':listSql[i]})
        i += 1
        print(make_data)
    return make_data